from tkinter import *
import tkinter as tk
from PIL import ImageTk, Image
import tkinter.messagebox
import mysql.connector
from tkinter import ttk
import sys
from tkinter import Menu
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment, PatternFill
from tkinter import messagebox
from tkinter import filedialog


def exit1():
    sys.exit()

def error_entry_name():
    tkinter.messagebox.showerror("Error", "Please enter a valid name")

def error_entry_phone():
    tkinter.messagebox.showerror("Error", "Please enter a valid phone")

def error_entry_email():
    tkinter.messagebox.showerror("Error", "Please enter a valid email")

def error_entry_gender():
    tkinter.messagebox.showerror("Error", "Please select a valid gender")

def error_entry_phone_n():
    tkinter.messagebox.showerror("Error", "Phone number should only contain numbers and 11 digits")

def error_entry_email_s():
    tkinter.messagebox.showerror("Error", "Email should contain @gmail.com or @example.com")

def duplicate_email():
    tkinter.messagebox.showerror("Error", "Email already exists")

def duplicate_phone():
    tkinter.messagebox.showerror("Error", "Phone number already exists")
    
def clear_listbox():
    frame111.destroy()  # Delete the frame111 widget
def error_entry_phones():
    tkinter.messagebox.showerror("Error", "Wrong entry: Name or phone number does not match", parent=status_window)
def error_modify_gender():
    tkinter.messagebox.showerror("Error", "Wrong entry: ID or gender does not match", parent=window2)

def update_status():
    global status_window,frame11
    global id1 
    id1 =StringVar()
    status_window=Toplevel(window)
    status_window.title("Employee database Status")
    status_window.geometry("600x700")
    status_window.maxsize(600, 700)
    status_window.config(bg="light sea green")
    Label(status_window, text="Welcome to Status Update!", font=("Century Schoolbook L", 20, "bold"), bg="deeppink").place(x=120, y=10)
    id_s = Label(status_window, text="ID:", font=("Arial", 15, "bold"), bg="light sea green")
    id_s.place(x=100, y=70)
    id_s = Entry(status_window, width=30, textvar=id1)
    id_s.place(x=170, y=70, height=30)
    frame11 = Frame(status_window, bg="lightcyan")
    frame11.place(x=100, y=120, height=150, width=400)


    def get_data_id():
        global frame111
        frame111=Frame(frame11, bg="lightgreen",height=100, width=400)
        frame111.pack()
        def check_frame():
            if frame111.winfo_exists():
                print("error")
        check_frame()
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="chair123$",
            database="mydatabase"
        )
        cursor = connection.cursor()
        query = "SELECT * FROM users where id=%s "
        values = (id1.get(),)
        cursor.execute(query,values)
        rows = cursor.fetchall()
        listbox1 = Listbox(frame111, font=("Arial", 12), bg="lightcyan", fg="darkblue",width=50,height=20)

        for row in rows:
            Id=f"ID: {row[0]}"
            name = f"Name: {row[1]}"
            phone = f"Phone Number: {row[2]}"
            email = f"Email: {row[3]}"
            gender = f"Gender: {row[4]}"
            Status = f"Status: {row[5]}"
            separator = "═════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════"

            listbox1.insert(END, Id)
            listbox1.insert(END, name)
            listbox1.insert(END, phone)
            listbox1.insert(END, email)
            listbox1.insert(END, gender)
            listbox1.insert(END, Status)
            listbox1.insert(END, separator)
        listbox1.pack(side=LEFT, fill=BOTH, expand=True)
    Status1 = Label(status_window, text="Status:", font=("Arial", 15, "bold"), bg="light sea green")
    Status1.place(x=100, y=410)
    check1 = Checkbutton(status_window, text="Active", variable=active_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check1.place(x=200, y=410)

    check2 = Checkbutton(status_window, text="Non Active", variable=non_active_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check2.place(x=300, y=410)

    check3 = Checkbutton(status_window, text="On Leave", variable=leave_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check3.place(x=200, y=450)

    check4 = Checkbutton(status_window, text="Fired BY", variable=fired_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check4.place(x=300, y=450)
    button1 = Button(status_window, text="Save", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=updated_status)
    button1.place(x=170, y=550)
    button2 = Button(status_window, text="Quit", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=exit1)
    button2.place(x=310, y=550)
    button3 = Button(status_window, text="Search", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("orange"), command=get_data_id)
    button3.place(x=170, y=300)
    button_clear = tk.Button(status_window, text="Clear", relief=tk.SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("orange"), command=clear_listbox)
    button_clear.place(x=310, y=300)

def updated_status():
    id1s=id1.get()
    status_c = ""
    if active_var.get() == 1:
        status_c += "Active"
    if non_active_var.get() == 1:
        status_c += "Non Active"
    if leave_var.get() == 1:
        status_c += "On Leave"
    if fired_var.get() == 1:
        status_c += "Fired BY"

    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    cursor = connection.cursor()
    query = "SELECT * FROM users WHERE id = %s "
    values = (id1s,)
    cursor.execute(query, values)
    row = cursor.fetchone()
    if row:
        query_update = "UPDATE users SET status = %s WHERE id = %s"
        values_update = (status_c,id1s,)
        cursor.execute(query_update, values_update)
        connection.commit()
        tkinter.messagebox.showinfo("Done", "Status updated successfully", parent=status_window)
        id1.set("")
        active_var.set(0)
        non_active_var.set(0)
        leave_var.set(0)
        fired_var.set(0)
        clear_listbox()
    else:
        error_entry_phones()
        # tkinter.messagebox.showerror("Error", "Wrong entry: Name or phone number does not match")
        # status_window.lift()  # Keep the status window in front
    cursor.close()
    connection.close()

def report_gen():
    query = "SELECT * FROM users WHERE status IN ("
    args = []
    if active_var.get() == 1:
        args.append("Active")
        query=query + "%s,"
    if non_active_var.get() == 1:
        args.append("Non Active")
        query=query + "%s,"
    if fired_var.get() == 1:
        args.append("Fired BY")
        query=query + "%s,"
    if leave_var.get() == 1:
        args.append("On Leave")
        query=query + "%s,"
    query = query[:len(query)-1]
    query=query + ")"
    # Create a new PDF file
    args=tuple(args)
    connection = mysql.connector.connect(host="localhost", user="root", password="chair123$", database="mydatabase")
    cursor = connection.cursor()
    cursor.execute(query,args)
    rows = cursor.fetchall()
    return rows


def generate_pdf_report():
    rows=report_gen()

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"report_{timestamp}.pdf"
    c = canvas.Canvas(filename, pagesize=letter)

    # Add your icon/image at the top
    image_path2 = "D:\\python program\\.vscode\\stylish-black-lion-logo-white-background-vector_532963-5993.jpg"
    c.drawImage(image_path2, 50, 700, width=60, height=60)  # Adjust the Y-coordinate as needed
    c.setFillColor(colors.red)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(60, 690, "RAA.CO")
    
    c.setFillColor(colors.red)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(50, 674, "RAA.CO: Copyright © [2023] by RAA.CO. All rights reserved.")
    # Draw a horizontal line
    c.line(50, 670, 560, 670)  # (x1, y1, x2, y2)
    c.line(50, 672, 560, 672)  # (x1, y1, x2, y2)
    # Add your content to the PDF
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(200, 690, "Employee Database Report")
    c.setFont("Helvetica", 10)
    c.drawString(450, 720, "Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    page_height = 792  # Height of a letter-sized page in points
    y = 650  # Starting Y-coordinate for the first row of data

    for row in rows:
        Id = f"ID: {row[0]}"
        name = f"Name: {row[1]}"
        phone = f"Phone Number: {row[2]}"
        email = f"Email: {row[3]}"
        gender = f"Gender: {row[4]}"
        Status = f"Status: {row[5]}"
        separator = "____________________________________________________________________________________________"

        # Example content
        c.setFont("Helvetica", 10)
        c.drawString(50, y, f"{Id}")
        c.drawString(50, y - 10, f"{name}")
        c.drawString(50, y - 20, f"{email}")
        c.drawString(50, y - 30, f"{phone}")
        c.drawString(50, y - 40, f"{gender}")
        c.drawString(50, y - 50, f"{Status}")
        c.drawString(50, y - 60, f"{separator}")
        c.drawString(50, y - 70," ")

        # Check if the next row fits on the current page
        y -= 80  # Adjust this value based on your desired spacing

        if y < 100 :
            c.line(50, 49, 560, 49)  # (x1, y1, x2, y2)
            c.line(50, 47, 560, 47)  # (x1, y1, x2, y2)
            c.setFillColor(colors.red)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(50, 51, " RAA.CO: This work is protected by copyright. Any unauthorized use or reproduction is prohibited by law.")
            c.showPage()  # Start a new page when the content exceeds the available space
            y = page_height - 50  # Reset Y-coordinate for the new page

    # Save the PDF
    c.save()



def generate_excel_report():
    # Create a new Excel workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Add data and formatting here
    sheet['A1'] = "Employee Database Report"
    sheet['A1'].font = Font(size=18, bold=True)
    sheet.merge_cells('A1:F1')  # Merge cells for the title
    sheet['A2'] = "Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    red_font = Font(color="FF0000")
    sheet['A8'] = f"Company Name: RAA.CO"
    sheet['A8'].font = Font(size=13,bold=True)
    sheet.merge_cells('A8:B8')  # Merge cells for the title
    # sheet['A3'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill color
    sheet['A8'].font = red_font
    sheet.merge_cells('A4:B4')

    # Set column widths for better visibility
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15

    # Add your icon/image to a specific cell (e.g., A5)
    # Replace 'image_path' with the actual path to your image
    # Make sure the image fits within the cell dimensions
    # You may need to adjust the row and column coordinates
    # Example:
    img = openpyxl.drawing.image.Image("D:\\python program\\.vscode\\stylish-black-lion-logo-white-background-vector_532963-5993.jpg")
    # img = Image("D:\\python program\\.vscode\\stylish-black-lion-logo-white-background-vector_532963-5993.jpg")
    img.width = 100  # Set the width of the image (adjust as needed)
    img.height = 100  # Set the height of the image (adjust as needed)
    sheet.add_image(img, 'A3')  # Add the image to cell 'C4'


    # Add header row
    headers = ["ID", "Name", "Email", "Phone Number", "Gender", "Status"]
    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=10, column=col_num)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Query your database and retrieve data into 'data' list
    # Ensure the 'data' list has rows of data in the same order as the headers

    # Retrieve data from MySQL database
    data=report_gen()

    # Populate data rows starting from row 6
    for row_num, row_data in enumerate(data,11):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the Excel file
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"report_{timestamp}.xlsx"
    wb.save(filename)

# You can call generate_excel_report() to generate the Excel report.


def report_window():
    report_window = Toplevel(window)
    report_window.title("Employee Database 2")
    report_window.geometry("600x500")
    report_window.maxsize(900, 700)
    report_window.config(bg="purple")
    Status1 = Label(report_window, text="Status:", font=("Arial", 20, "bold"), bg="light sea green")
    Status1.place(x=80, y=115)
    check1 = Checkbutton(report_window, text="Active", variable=active_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check1.place(x=200, y=80)

    check2 = Checkbutton(report_window, text="Non Active", variable=non_active_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check2.place(x=350, y=80)

    check3 = Checkbutton(report_window, text="On Leave", variable=leave_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check3.place(x=200, y=160)

    check4 = Checkbutton(report_window, text="Fired BY", variable=fired_var, font=("Arial", 13, "bold"), bg="light sea green", onvalue=1, offvalue=0)
    check4.place(x=350, y=160)

    button1 = Button(report_window, text="Report in PDF", relief=SOLID, width=15, height=1, font=("arial", 15, "bold"), bg=("deeppink"),command=generate_pdf_report)
    button1.place(x=110, y=250)
    button2 = Button(report_window, text="Report in EXCEL ", relief=SOLID, width=15, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=generate_excel_report)
    button2.place(x=310, y=250)

def security():
    global photo5
    # Create a MySQL connection
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    def signup():
        username = signup_username.get()
        password = signup_password.get()

        # Check if the username already exists
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM users_pass WHERE username = %s", (username,))
        existing_user = cursor.fetchone()

        if existing_user:
            tkinter.messagebox.showerror("Error", "Username already exists. Please choose a different one.",parent=window3)
        else:
            # cursor.execute("INSERT INTO users_pass (username, password) VALUES (%s, %s)", (username, password))
            # connection.commit()
            tkinter.messagebox.showinfo("Sorry", "Signup Account Full. Contact admin to grant acess. You can now contact on 03135749021.",parent=window3)
            signup_username.delete(0, 'end')  # Clear the username entry
            signup_password.delete(0, 'end')  # Clear the password entry

    def login():
        username = login_username.get()
        password = login_password.get()
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM users_pass WHERE username = %s AND password = %s", (username, password))
        user = cursor.fetchone()

        if user:
            tkinter.messagebox.showinfo("Success", "Login successful.",parent=window3)
            modify_user_information()
            login_username.delete(0, 'end')  # Clear the username entry
            login_password.delete(0, 'end')  # Clear the password entry
        else:
            tkinter.messagebox.showerror("Error", "Invalid username or password.",parent=window3)

    # Create a default admin account
    def create_default_admin():
        tkinter.messagebox.showwarning("Reminder", "Username = RAABBASI And Password =123.",parent=window3)
    window3 = Toplevel(window)
    window3.title("Security Check")
    window3.geometry("1000x700")
    window3.maxsize(700, 500)
    image1 = Image.open("D:\\FYP project Python Desktop App\\.vscode\\pback.jpg")
    image1 = image1.resize((700, 500))
    photo5 = ImageTk.PhotoImage(image1)
    Label(window3, image=photo5).place(x=0, y=0)
    signup_frame = tk.Frame(window3,bg="light sea green",width=350 ,height=250)
    signup_frame.pack(pady=20)

    signup_label = tk.Label(signup_frame, text="Signup" ,font=("Arial", 15, "bold"), bg="light sea green",)
    signup_label.grid(row=0, column=0, columnspan=2)

    signup_username_label = tk.Label(signup_frame, text="Username:",font=("Arial", 15, "bold"), bg="light sea green")
    signup_username_label.grid(row=1, column=0)
    signup_username = tk.Entry(signup_frame)
    signup_username.grid(row=1, column=1)

    signup_password_label = tk.Label(signup_frame, text="Password:",font=("Arial", 15, "bold"), bg="light sea green")
    signup_password_label.grid(row=2, column=0)
    signup_password = tk.Entry(signup_frame, show="*")
    signup_password.grid(row=2, column=1)

    signup_button = tk.Button(signup_frame, text="Signup",font=("Arial", 15, "bold"), bg="light sea green", command=signup)
    signup_button.grid(row=3, column=0, columnspan=2)

    # Create and configure the login frame
    login_frame = tk.Frame(window3,bg="light sea green",width=350,height=250)
    login_frame.pack(pady=20)

    login_label = tk.Label(login_frame, text="Login",font=("Arial", 15, "bold"), bg="light sea green")
    login_label.grid(row=0, column=0, columnspan=2)

    login_username_label = tk.Label(login_frame, text="Username:",font=("Arial", 15, "bold"), bg="light sea green")
    login_username_label.grid(row=1, column=0)
    login_username = tk.Entry(login_frame)
    login_username.grid(row=1, column=1)

    login_password_label = tk.Label(login_frame, text="Password:",font=("Arial", 15, "bold"), bg="light sea green")
    login_password_label.grid(row=2, column=0)
    login_password = tk.Entry(login_frame, show="*")
    login_password.grid(row=2, column=1)

    login_button = tk.Button(login_frame, text="Login",font=("Arial", 15, "bold"), bg="light sea green", command=login)
    login_button.grid(row=3, column=0, columnspan=2)

    login_button = tk.Button(window3, text="Default Account", font=("Arial", 15, "bold"),width=20, bg="light sea green",command=create_default_admin)
    login_button.place(x=220,y=400)

    window3.mainloop()

def modify_user_information():
    global photo4,window2,id_m,name_m,email_m,phone_m,gender_m
    id_m=StringVar()
    gender_m=StringVar()
    gender_m.set(None)
    name_m=StringVar()
    email_m=StringVar()
    phone_m=StringVar()
    window2 = Toplevel(window)
    window2.title("Employee database 1")
    window2.geometry("1000x700")
    window2.maxsize(1000, 700)
    image1 = Image.open("D:\\FYP project Python Desktop App\\.vscode\\pback.jpg")
    image1 = image1.resize((1000, 700))
    photo4 = ImageTk.PhotoImage(image1)
    Label(window2, image=photo4).place(x=0, y=0)
    Label(window2, text="what do you want to modify name, phone_number, email, and gender", font=("Century Schoolbook L", 17, "bold"), bg="deeppink").place(x=100, y=10)
    user_id_label = Label(window2, text="User ID:", font=("Arial", 15, "bold"), bg="light sea green")
    user_id_label.place(x=250, y=270)
    name1 = Label(window2, text="Name:", font=("Arial", 15, "bold"), bg="light sea green")
    name1.place(x=250, y=330)
    email1 = Label(window2, text="Email:", font=("Arial", 15, "bold"), bg="light sea green")
    email1.place(x=250, y=390)
    phone_no1 = Label(window2, text="Phone no:", font=("Arial", 15, "bold"), bg="light sea green")
    phone_no1.place(x=250, y=450)
    gender1 = Label(window2, text="Gender:", font=("Arial", 15, "bold"), bg="light sea green")
    gender1.place(x=250, y=510)
    check1 = Radiobutton(window2, text="Male", variable=gender_m, value="Male", relief=FLAT, width=5, height=-1, font=("arial", 13, "bold"), bg=("light sea green"))
    check1.place(x=370, y=510)

    check2 = Radiobutton(window2, text="Female", variable=gender_m, value="Female", relief=FLAT, width=6, height=-1, font=("arial", 13, "bold"), bg=("light sea green"))
    check2.place(x=460, y=510)
    user_id_entry = Entry(window2, width=30, textvar=id_m)
    user_id_entry.place(x=370, y=270, height=30)
    name_entry = Entry(window2, width=30, textvar=name_m)
    name_entry.place(x=370, y=330, height=30)
    email_entry = Entry(window2, width=30, textvar=email_m)
    email_entry.place(x=370, y=390, height=30)
    phone_entry = Entry(window2, width=30, textvar=phone_m)
    phone_entry.place(x=370, y=450, height=30)
    modify_button = Button(window2, text="Save", font=("Arial", 15, "bold"), bg="light sea green", command=modify)
    modify_button.place(x=400, y=570)
def modify():
    if gender_m.get()=="" and id_m.get().isdigit() and email_m.get()=="" and phone_m.get()=="":#name
        query_update = "UPDATE users SET name = %s WHERE id = %s"
        values_update = (name_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if name_m.get()=="" and id_m.get().isdigit() and gender_m.get()=="" and phone_m.get()=="":#email
        query_update = "UPDATE users SET email = %s WHERE id = %s"
        values_update = (email_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if name_m.get()=="" and id_m.get().isdigit() and email_m.get()=="" and gender_m.get()=="":#phone
        query_update = "UPDATE users SET phone_number = %s WHERE id = %s"
        values_update = (phone_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and email_m.get()=="" and phone_m.get()=="":#name,gender
        query_update = "UPDATE users SET name = %s, gender = %s  WHERE id = %s"
        values_update = (name_m.get(),gender_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and email_m.get()=="" and name_m.get()=="":#phone,gender
        query_update = "UPDATE users SET phone_number = %s, gender = %s  WHERE id = %s"
        values_update = (phone_m.get(),gender_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and name_m.get()=="" and phone_m.get()=="":#email,gender
        query_update = "UPDATE users SET email = %s, gender = %s  WHERE id = %s"
        values_update = (email_m.get(),gender_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and email_m.get()=="" and gender_m.get()=="":#name,phone
        query_update = "UPDATE users SET name = %s, phone_number = %s  WHERE id = %s"
        values_update = (name_m.get(),phone_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and gender_m.get()=="" and phone_m.get()=="":#name,email
        query_update = "UPDATE users SET name = %s, email = %s  WHERE id = %s"
        values_update = (name_m.get(),email_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and name_m.get()=="" and gender_m.get()=="":#phone,email
        query_update = "UPDATE users SET phone_number = %s, email = %s  WHERE id = %s"
        values_update = (phone_m.get(),email_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and phone_m.get()=="":#name,gender,email
        query_update = "UPDATE users SET name = %s, gender = %s, email=%s  WHERE id = %s"
        values_update = (name_m.get(),gender_m.get(),email_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and name_m.get()=="":#phone,gender,email
        query_update = "UPDATE users SET phone_number = %s, gender = %s, email = %s  WHERE id = %s"
        values_update = (phone_m.get(),gender_m.get(),email_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and email_m.get()=="":#name,gender,phone
        query_update = "UPDATE users SET name = %s, gender = %s, phone_number = %s  WHERE id = %s"
        values_update = (name_m.get(),gender_m.get(),phone_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit() and gender_m.get()=="":#name,email,phone
        query_update = "UPDATE users SET name = %s, email = %s, phone_number = %s  WHERE id = %s"
        values_update = (name_m.get(),email_m.get(),phone_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if id_m.get().isdigit():#name,gender,email,phone
        query_update = "UPDATE users SET name = %s, gender = %s, email=%s ,phone_number = %s  WHERE id = %s"
        values_update = (name_m.get(),gender_m.get(),email_m.get(),phone_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

    if name_m.get()=="" and id_m.get().isdigit() and email_m.get()=="" and phone_m.get()=="":#gender
        query_update = "UPDATE users SET gender = %s WHERE id = %s"
        values_update = (gender_m.get(),id_m.get(),)
        modify_update(query_update,values_update)

def modify_update(query_update,values_update):
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    cursor = connection.cursor()
    query = "SELECT * FROM users WHERE id = %s "
    values = (id_m.get(),)
    cursor.execute(query, values)
    row = cursor.fetchone()
    if row:
        cursor.execute(query_update, values_update)
        connection.commit()
        tkinter.messagebox.showinfo("Done", "Updated successfully", parent=window2)
        id_m.set("")
        gender_m.set(None)
        email_m.set("")
        phone_m.set("")
        name_m.set("")

    else:
        error_entry_phones()
        # tkinter.messagebox.showerror("Error", "Wrong entry: Name or phone number does not match")
        # status_window.lift()  # Keep the status window in front
    cursor.close()
    connection.close()

def check():
    tkinter.messagebox.showinfo("Done", "No Updates are avilable")
def full_screen():
    window.state('zoomed')
def help():
    tkinter.messagebox.showinfo("Done", """To get a list of available modules, keywords, symbols, or topics, type
    "modules", "keywords", "symbols", or "topics".  Each module also comes
    with a one-line summary of what it does; to list the modules whose name
    or summary contain a given string such as "spam", type "modules spam".""")
def short_screen():
    window.geometry("600x00")

def normal_screen():
    window.geometry("900x700")

def file_new():
    print("New File")

def file_open():
    print("Open File")

def file_exit():
    window.destroy()

# Connect to the database (You may need to adjust the connection parameters)
def get_database_connection():
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    return connection


def search_suggestions(search_term,query):
    suggestions_list = []
    
    connection = get_database_connection()
    cursor = connection.cursor()
    
    cursor.execute(query, ("%" + search_term + "%",))
    rows = cursor.fetchall()
    
    for row in rows:
        suggestions_list.append(row[0])
    
    cursor.close()
    connection.close()
    
    return suggestions_list

def byname(search_by):
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    if search_by =="Name":
        query = "SELECT id, name , gender, phone_number,email FROM users WHERE name = %s"
    elif search_by =="Email":
        query = "SELECT id, name, gender, phone_number,email FROM users WHERE email = %s"
    else:
        query = "SELECT id, name, gender, phone_number,email FROM users WHERE phone_number = %s"

    cursor = connection.cursor()
    cursor.execute(query, (name2.get(),))
    rows = cursor.fetchall()
    display_window_search1 = Toplevel(window)
    display_window_search1.title(f"Searching by {search_by}")
    display_window_search1.geometry("800x600")
    display_window_search1.maxsize(800, 600)
    frame_results = Frame(display_window_search1, width=800, height=500, bg="lightcyan")
    frame_results.pack()
    scrollbar = Scrollbar(frame_results)
    scrollbar.pack(side=RIGHT, fill=Y)
    listbox = Listbox(frame_results, font=("Arial", 15), width=800, bg="lightcyan", fg="darkblue", yscrollcommand=scrollbar.set)
    for row in rows:
        Id=f"ID: {row[0]}"
        name = f"Name: {row[1]}"
        gender = f"Gender: {row[2]}"
        phone = f"Phone: {row[3]}"
        email = f"Phone: {row[4]}"
        separator = "═════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════"

        listbox.insert(END, Id)
        listbox.insert(END, name)
        listbox.insert(END,phone)
        listbox.insert(END, gender)
        listbox.insert(END, email)
        listbox.insert(END, separator)

    scrollbar.config(command=listbox.yview)
    listbox.pack(side=LEFT, fill=BOTH, expand=True)
    cursor.close()
    connection.close()
    name2.set("")

def Search_window(search_by):
    def close_window():
        search_window_all.destroy()  # Close the window
    def autocomplete(event):
        if search_by =="Name":
            search_term = all_search.get()
            query = "SELECT name FROM users WHERE name LIKE %s"
            suggestions = search_suggestions(search_term,query)
            all_search['values'] = suggestions
        elif search_by =="Email":
            search_term=all_search.get()
            query = "SELECT email FROM users WHERE email LIKE %s"
            suggestions=search_suggestions(search_term,query)
            all_search['values']=suggestions
        else:
            query = "SELECT phone_number FROM users WHERE phone_number LIKE %s"
            search_term = all_search.get()
            suggestions = search_suggestions(search_term,query)
            all_search['values'] = suggestions


    search_window_all = Toplevel(window)
    search_window_all.title(f"Searching by {search_by}")
    search_window_all.geometry("800x600")
    search_window_all.maxsize(800, 600)
    frame_by_name = Frame(search_window_all, width=800, height=50, bg="cyan")
    frame_by_name.grid(row=1, column=0, padx=0, pady=10)
    label_search_1 = Label(frame_by_name, text=f"{search_by} to search", bg="cyan", font=("arial", 28, "bold"))
    label_search_1.grid(row=0, column=0)
    all_search = ttk.Combobox(frame_by_name, width=60, textvar=name2)
    all_search.grid(row=0, column=1, padx=120, pady=10)
    all_search.bind("<KeyRelease>", autocomplete)  # Bind the autocomplete function to the key release event
    frame_by_name_3 = Frame(search_window_all, width=800, height=50, bg="cyan")
    frame_by_name_3.grid(row=2, column=0)
    frame_by_name_2 = Frame(search_window_all, width=800, height=50, bg="cyan")
    frame_by_name_2.grid(row=5, column=0)
    button3_display_window = tk.Button(frame_by_name_2, text="Search", relief=tk.SOLID, width=15, height=1, font=("arial", 15, "bold"), bg="deeppink",command=lambda: [close_window(),byname(search_by)])
    button3_display_window.grid(row=3, column=3, padx=10, pady=10)
    button2 =tk.Button(frame_by_name_2, text="Quit", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=close_window)
    button2.grid(row=3,column=1)

def display_information():
    def close_window():
        display_window.destroy()  # Close the window
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    cursor = connection.cursor()
    query = "SELECT * FROM users"
    cursor.execute(query)
    rows = cursor.fetchall()

    display_window = Toplevel(window)
    display_window.title("Employee Database 2")
    display_window.geometry("900x700")
    display_window.maxsize(900, 700)

    frame1 = Frame(display_window, bg="lightcyan")
    frame1.place(x=0, y=0, height=500, width=900)

    frame2 = Frame(display_window, bg="blue")
    frame2.place(x=0, y=500, width=900, height=200)

    Label(frame2, text="Searching Options", font=("arial", 18, "bold"), bg="deep pink").grid(row=1, column=1)

    scrollbar = Scrollbar(frame1)
    scrollbar.pack(side=RIGHT, fill=Y)

    listbox = Listbox(frame1, font=("Arial", 12), bg="lightcyan", fg="darkblue", yscrollcommand=scrollbar.set)

    for row in rows:
        Id=f"ID: {row[0]}"
        name = f"Name: {row[1]}"
        phone = f"Phone Number: {row[2]}"
        email = f"Email: {row[3]}"
        gender = f"Gender: {row[4]}"
        Status = f"Status: {row[5]}"
        separator = "═════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════"

        listbox.insert(END, Id)
        listbox.insert(END, name)
        listbox.insert(END, phone)
        listbox.insert(END, email)
        listbox.insert(END, gender)
        listbox.insert(END, Status)
        listbox.insert(END, separator)

    scrollbar.config(command=listbox.yview)
    listbox.pack(side=LEFT, fill=BOTH, expand=True)

    button1_display_window = tk.Button(frame2, text="BY NAME", relief=tk.SOLID, width=10, height=1,font=("arial", 15, "bold"), bg="deeppink", command=lambda: [close_window(),Search_window("Name")])
    button1_display_window.grid(row=2, column=0)

    button2_display_window = tk.Button(frame2, text="BY EMAIL", relief=tk.SOLID, width=10, height=1,font=("arial", 15, "bold"), bg="deeppink", command=lambda: [close_window(),Search_window("Email")])
    button2_display_window.grid(row=2, column=1, padx=10, pady=10)

    button3_display_window = tk.Button(frame2, text="BY PHONE NO", relief=tk.SOLID, width=15, height=1,font=("arial", 15, "bold"), bg="deeppink",command=lambda: [close_window(),Search_window("Phone")])
    button3_display_window.grid(row=2, column=2, padx=10, pady=10)
    button2 =tk.Button(frame2, text="Quit", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=close_window)
    button2.grid(row=3,column=1)
    modify_button = Button(frame2, text="Modify Information", relief=SOLID, width=16, height=1, font=("arial", 15, "bold"), bg=("deeppink"),command=lambda: [close_window(),security()])
    modify_button.grid(row=2,column=3)
    modify_button = Button(frame2, text="Reports", relief=SOLID, width=16, height=1, font=("arial", 15, "bold"), bg=("deeppink"),command=lambda: [close_window(),report_window()])
    modify_button.grid(row=3,column=2)

    cursor.close()
    connection.close()
    display_window.mainloop()

def store_information(name, phone_number, email, gender,status):
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="chair123$",
        database="mydatabase"
    )
    cursor = connection.cursor()

    query_email = "SELECT * FROM users WHERE email = %s"
    cursor.execute(query_email, (email,))
    if cursor.fetchone():
        duplicate_email()
        cursor.fetchall()
        cursor.close()
        connection.close()
        return

    query_phone = "SELECT * FROM users WHERE phone_number = %s"
    cursor.execute(query_phone, (phone_number,))
    if cursor.fetchone():
        duplicate_phone()
        cursor.fetchall()
        cursor.close()
        connection.close()
        return

    query = "INSERT INTO users (name, phone_number, email, gender,status) VALUES (%s, %s, %s, %s,%s)"
    values = (name, phone_number, email, gender,status)
    cursor.execute(query, values)
    connection.commit()

    cursor.close()
    connection.close()

    tkinter.messagebox.showinfo("Done", "Information stored successfully")
    name_g.set("")
    email_g.set("")
    phone_g.set("")
    gender_g.set(None)

def database():
    name = name_g.get()
    phone_number = phone_g.get()
    email = email_g.get()
    gender = gender_g.get()
    status = "Active"  # Assuming the default status is active

    if name == '':
        error_entry_name()
        return
    if email == '' or "@" not in email:
        error_entry_email()
        return
    if len(phone_number) != 11 or not phone_number.isdigit():
        error_entry_phone()
        return
    if gender == "None":
        error_entry_gender()
        return

    store_information(name, phone_number, email, gender,status)  # Existing store_information function


window = Tk()
window.title("Employee database 1")
window.geometry("900x700")
window.maxsize(900, 700)

image1 = Image.open("D:\\FYP project Python Desktop App\\.vscode\\pback.jpg")
image1 = image1.resize((900, 700))
photo1 = ImageTk.PhotoImage(image1)
Label(window, image=photo1).place(x=0, y=0)

image2 = Image.open("D:\\FYP project Python Desktop App\\.vscode\\database.png")
image2 = image2.resize((200, 200))
photo2 = ImageTk.PhotoImage(image2)
Label(window, image=photo2).place(x=350, y=80)

# create a menu bar
menu_bar = Menu(window)
# # create a File menu
file_menu = Menu(menu_bar, tearoff=0)
file_menu.add_command(label="New")
file_menu.add_command(label="Open")
file_menu.add_separator()
file_menu.add_command(label="Exit", command=exit1)
# add the File menu to the menu bar
menu_bar.add_cascade(label="File", menu=file_menu)


view_menu=Menu(menu_bar, tearoff=0)
view_menu.add_command(label="Full screen",command=full_screen)
view_menu.add_command(label="Short screen",command=short_screen)
view_menu.add_command(label="Normal screen")
menu_bar.add_cascade(label="View",menu=view_menu)

help_menu=Menu(menu_bar,tearoff=0)
help_menu.add_command(label="check for updates",command=check)
help_menu.add_command(label="help",command=help)
menu_bar.add_cascade(label="Help",menu=help_menu)

name_g = StringVar()
email_g = StringVar()
phone_g = StringVar()
gender_g = StringVar()
name2 = StringVar()
email2 = StringVar()
phone2 = StringVar()
new_name_g=StringVar()
active_var = IntVar(value=0)
non_active_var = IntVar(value=0)
leave_var = IntVar(value=0)
fired_var = IntVar(value=0)
gender_g.set(None)
name_list=[]
phone_list=[]
email_list=[]

Label(window, text="Welcome to Employee Database!", font=("Century Schoolbook L", 25, "bold"), bg="deeppink").place(x=200, y=20)

name1 = Label(window, text="Name:", font=("Arial", 15, "bold"), bg="light sea green")
name1.place(x=250, y=330)

email1 = Label(window, text="Email:", font=("Arial", 15, "bold"), bg="light sea green")
email1.place(x=250, y=390)

phone_no1 = Label(window, text="Phone no:", font=("Arial", 15, "bold"), bg="light sea green")
phone_no1.place(x=250, y=450)

gender1 = Label(window, text="Gender:", font=("Arial", 15, "bold"), bg="light sea green")
gender1.place(x=250, y=510)

name_entry = Entry(window, width=30, textvar=name_g)
name_entry.place(x=370, y=330, height=30)

email_entry = Entry(window, width=30, textvar=email_g)
email_entry.place(x=370, y=390, height=30)

phone_entry = Entry(window, width=30, textvar=phone_g)
phone_entry.place(x=370, y=450, height=30)

button1 = Button(window, text="Save", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=database)
button1.place(x=300, y=550)

button2 = Button(window, text="Quit", relief=SOLID, width=10, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=exit1)
button2.place(x=450, y=550)

check1 = Radiobutton(window, text="Male", variable=gender_g, value="Male", relief=FLAT, width=5, height=-1, font=("arial", 13, "bold"), bg=("light sea green"))
check1.place(x=370, y=510)

check2 = Radiobutton(window, text="Female", variable=gender_g, value="Female", relief=FLAT, width=6, height=-1, font=("arial", 13, "bold"), bg=("light sea green"))
check2.place(x=470, y=510)

button3 = Button(window, text="Display Information", relief=SOLID, width=16, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=display_information)
button3.place(x=230, y=610)

modify_button = Button(window, text="Status Update", relief=SOLID, width=16, height=1, font=("arial", 15, "bold"), bg=("deeppink"), command=update_status)
modify_button.place(x=450, y=610)

# add the menu bar to the window
window.config(menu=menu_bar)
window.mainloop()
